#!/usr/bin/env python3
"""
Automated Data Synchronizer for Bosch Dashboards:
1. Program Performance Dashboard (index.html, data.json from 'vm_rank_report')
2. QC Performance Dashboard (qc.html, data_qc.json from 'QC Tracker')
Downloads latest live datasets directly from SharePoint export URLs,
embeds parsed structured datasets into index.html and qc.html,
and optionally pushes updates to GitHub.
"""

import urllib.request
import http.cookiejar
from pyxlsb import open_workbook
import openpyxl
import io
import json
import datetime
import os
import subprocess
import sys

SHAREPOINT_PROGRAM_URL = "https://teamchannelplay-my.sharepoint.com/:x:/g/personal/bikash_roy1_channelplay_in/IQC5rvkJBy1dRYzVkAjdHFo6AaOSY7IoFmC7EVtBuJZzCAA?e=u5PYBz&download=1"
SHAREPOINT_QC_URL = "https://teamchannelplay-my.sharepoint.com/:x:/g/personal/bikash_roy1_channelplay_in/IQDvQ_YpN5ZKQoyoPYtSEiPeAQXU4oOQqunN3wp0jnjdBX0?e=IiKb9c&download=1"

PROGRAM_LOCAL_FALLBACKS = [
    "/Users/bikash/Library/CloudStorage/OneDrive-ChannelplayLimited/My Laptop/0 Active Projects/Bosch VM/Dashboard Data/Downloads/Bosch AOM's Report Aug'26.xlsb",
    "/Users/bikash/Downloads/Bosch AOM's Report Aug'26.xlsb",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "Bosch AOM's Report Aug'26.xlsb")
]

QC_LOCAL_FALLBACKS = [
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "qc_download.xlsx"),
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "QC_Tracker.xlsx")
]

def get_opener():
    cookie_jar = http.cookiejar.CookieJar()
    return urllib.request.build_opener(
        urllib.request.HTTPCookieProcessor(cookie_jar),
        urllib.request.HTTPRedirectHandler
    )

def fetch_url(url, fallback_paths, label="dataset"):
    print(f"Fetching latest {label} from SharePoint...")
    opener = get_opener()
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}
    )
    try:
        data = opener.open(req, timeout=45).read()
        if len(data) > 3000:
            print(f"✓ Downloaded {len(data):,} bytes for {label}")
            return io.BytesIO(data)
    except Exception as e:
        print(f"SharePoint fetch failed for {label}: {e}. Checking local fallbacks...")

    for path in fallback_paths:
        if os.path.exists(path):
            print(f"✓ Using local fallback for {label}: {path}")
            with open(path, "rb") as f:
                return io.BytesIO(f.read())

    print(f"⚠️ Could not download {label} or locate local fallback file.")
    return None

# ----------------- 1. PROGRAM DATA PARSING -----------------
def parse_program_dataset(wb_stream):
    with open_workbook(wb_stream) as wb:
        sheet_name = "vm_rank_report" if "vm_rank_report" in wb.sheets else wb.sheets[0]
        print(f"Reading sheet: {sheet_name}")
        with wb.get_sheet(sheet_name) as s:
            rows = list(s.rows())

    if len(rows) < 2:
        raise ValueError("Sheet does not contain enough rows.")

    headers = [cell.v for cell in rows[1][:39]]
    vm_records = []

    for r in rows[2:]:
        vals = [cell.v for cell in r[:39]]
        if len(vals) > 3 and vals[0] and vals[3]:
            rec = {}
            for idx, h in enumerate(headers):
                if h:
                    v = vals[idx] if idx < len(vals) else None
                    rec[h] = v
                    if "unproductive visit" in str(h).lower():
                        if "%" in str(h):
                            rec["Productive Visit %"] = v
                        else:
                            rec["Productive Visits"] = v
            vm_records.append(rec)

    aom_base_towns = {
        'Irfan Ali': 'Hyderabad',
        'Najeeb Hydrose': 'Cochin',
        'Pratik Kumar  Pandey': 'Delhi',
        'Pratik Kumar Pandey': 'Delhi',
        'Danish Khan': 'Mumbai'
    }
    aom_dict = {}
    for r in vm_records:
        aom = r.get('AOM Name')
        if not aom:
            continue
        if aom not in aom_dict:
            aom_dict[aom] = {
                'Aom Name': aom,
                'Base Town': aom_base_towns.get(aom, r.get('Base Town', '--')),
                'cov_tgt': 0, 'cov_ach': 0,
                'cata_tgt': 0, 'cata_ach': 0,
                'md_tgt': 0, 'md_ach': 0,
                'visits_ach': 0,
                'adh_sum': 0,
                'final_sum': 0,
                'count': 0
            }
        acc = aom_dict[aom]
        acc['count'] += 1
        acc['cov_tgt'] += float(r.get('Coverage Target') or 0)
        acc['cov_ach'] += float(r.get('Coverage Achievement') or 0)
        acc['cata_tgt'] += float(r.get('Cat A Target') or 0)
        acc['cata_ach'] += float(r.get('Cat A 1st visit') or 0)
        acc['md_tgt'] += float(r.get('Target Man-days') or 0)
        acc['md_ach'] += float(r.get('Man-days Achieved') or 0)
        acc['visits_ach'] += float(r.get('Visits Achieved') or 0)
        acc['adh_sum'] += float(r.get('First Visit Time Adherence (11:30 AM)') or 0)
        acc['final_sum'] += float(r.get('Final Achievement') or 0)

    aom_records = []
    for aom, acc in aom_dict.items():
        n = acc['count'] or 1
        cov_pct = acc['cov_ach'] / acc['cov_tgt'] if acc['cov_tgt'] else 0
        cata_pct = acc['cata_ach'] / acc['cata_tgt'] if acc['cata_tgt'] else 0
        md_pct = acc['md_ach'] / acc['md_tgt'] if acc['md_tgt'] else 0
        prod = acc['visits_ach'] / acc['md_ach'] if acc['md_ach'] else 0
        adh = acc['adh_sum'] / n
        final_ach = acc['final_sum'] / n
        aom_records.append({
            'Aom Name': aom,
            'Base Town': acc['Base Town'],
            'Coverage %': cov_pct,
            'Cat A Coverage': cata_pct,
            'Achi Mandays %': md_pct,
            'Productivity': prod,
            'First Visit Time Adherence (11:30 AM)': adh,
            'Final Achievement': final_ach
        })

    aom_records.sort(key=lambda x: x['Final Achievement'], reverse=True)
    for idx, item in enumerate(aom_records):
        item['Rank'] = idx + 1

    return {
        "lastSynced": datetime.datetime.now().strftime("%d %b %Y, %I:%M %p"),
        "timestamp": datetime.datetime.now().isoformat(),
        "totalVMs": len(vm_records),
        "totalAOMs": len(aom_records),
        "vms": vm_records,
        "aoms": aom_records
    }

# ----------------- 2. QC DATA PARSING -----------------
def parse_qc_dataset(wb_stream):
    wb = openpyxl.load_workbook(wb_stream, data_only=True)
    target_sheet_name = None
    for name in wb.sheetnames:
        if "qc tracker" in name.lower() or "qc" in name.lower():
            target_sheet_name = name
            break
    if not target_sheet_name:
        target_sheet_name = wb.sheetnames[0]

    print(f"Reading QC sheet: {target_sheet_name}")
    sheet = wb[target_sheet_name]

    def serialize_val(v):
        if isinstance(v, (datetime.date, datetime.datetime)):
            return v.strftime('%Y-%m-%d')
        return v

    rows = []
    for r in range(1, sheet.max_row + 1):
        row_vals = [serialize_val(sheet.cell(r, c).value) for c in range(1, sheet.max_column + 1)]
        if any(x is not None for x in row_vals):
            rows.append(row_vals)

    return rows

# ----------------- 3. HTML EMBEDDING -----------------
def embed_into_html(data, html_file, json_file=None):
    compact_json = json.dumps(data, separators=(",", ":"))

    if json_file:
        with open(json_file, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        print(f"✓ Saved {json_file}")

    if not os.path.exists(html_file):
        print(f"⚠️ HTML file {html_file} does not exist.")
        return

    with open(html_file, "r", encoding="utf-8") as f:
        html = f.read()

    tag_start = '<script id="sample-data" type="application/json">'
    tag_end = '</script>'

    idx1 = html.find(tag_start)
    if idx1 != -1:
        idx2 = html.find(tag_end, idx1)
        if idx2 != -1:
            html = html[:idx1 + len(tag_start)] + compact_json + html[idx2:]
            with open(html_file, "w", encoding="utf-8") as f:
                f.write(html)
            print(f"✓ Embedded live data into {html_file}")
            return

    print(f"⚠️ Could not find sample-data tag in {html_file}")

def auto_push_to_github():
    print("\n--- Checking for GitHub updates ---")
    try:
        files = ["index.html", "data.json", "qc.html", "data_qc.json"]
        subprocess.run(["git", "add"] + [f for f in files if os.path.exists(f)], check=True)
        res = subprocess.run(["git", "diff", "--staged", "--quiet"])
        if res.returncode != 0:
            msg = f"chore(data): auto-sync latest live SharePoint datasets ({datetime.datetime.now().strftime('%Y-%m-%d %H:%M')})"
            subprocess.run(["git", "commit", "-m", msg], check=True)
            subprocess.run(["git", "push", "origin", "main"], check=True)
            print("🚀 Successfully pushed updated datasets to GitHub main branch!")
        else:
            print("No data changes detected. Remote repository is already up-to-date.")
    except Exception as e:
        print("Note: Could not push to git automatically:", e)

def main():
    print("=== Bosch VM Dashboards: Dual Data Synchronizer ===")

    # 1. Sync Program Performance Data
    try:
        stream_prog = fetch_url(SHAREPOINT_PROGRAM_URL, PROGRAM_LOCAL_FALLBACKS, "Program Report (xlsb)")
        if stream_prog:
            dataset_prog = parse_program_dataset(stream_prog)
            print(f"Successfully parsed Program Data: {dataset_prog['totalVMs']} VMs, {dataset_prog['totalAOMs']} AOMs")
            embed_into_html(dataset_prog, "index.html", "data.json")
    except Exception as e:
        print("Error syncing Program Performance Data:", e)

    # 2. Sync QC Performance Data
    try:
        stream_qc = fetch_url(SHAREPOINT_QC_URL, QC_LOCAL_FALLBACKS, "QC Tracker (xlsx)")
        if stream_qc:
            dataset_qc = parse_qc_dataset(stream_qc)
            print(f"Successfully parsed QC Tracker Data: {len(dataset_qc)} rows")
            embed_into_html(dataset_qc, "qc.html", "data_qc.json")
    except Exception as e:
        print("Error syncing QC Performance Data:", e)

    if "--push" in sys.argv:
        auto_push_to_github()

if __name__ == "__main__":
    main()
